# -*- coding: utf-8 -*-
"""知识层 + 设计 + 文档自动化 的回归测试。"""
import os, sys, pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from envcad.knowledge import materials, codes, theory, materials_summary, code_summary
from envcad.design.rc_beam import design_rc_beam, format_rc_beam_result
from envcad.design.section_db import select_section, format_section_choice
from envcad.docgen.spec_doc import generate_structure_spec
from envcad.docgen.calc_book import generate_calc_book
from envcad.docgen.bom_xlsx import generate_material_bom


def test_knowledge_imports():
    assert len(materials.CONCRETE) >= 14
    assert len(materials.REBAR_GRADE) >= 4
    assert len(codes.GB_CODES) >= 10
    assert len(theory.all_principles()) >= 10
    assert "混凝土" in materials_summary()
    assert "GB" in code_summary()


def test_rc_beam_design():
    r = design_rc_beam(250, 500, 20, "C30", "HRB400",
                       M=120e6, V=180e3, l=6000)
    assert r["rebar"]["area"] > 0
    assert r["all_ok"] is True, format_rc_beam_result(r)


def test_section_select():
    ch = select_section("I", 982.0)
    assert ch["Ax"] >= 982.0
    assert "I" in ch["name"]


def test_spec_doc(tmp_path):
    p = os.path.join(tmp_path, "说明.docx")
    generate_structure_spec(p, project="测试工程")
    assert os.path.getsize(p) > 0
    from docx import Document
    doc = Document(p)
    assert any("设计依据" in (par.text or "") for par in doc.paragraphs)


def test_calc_book(tmp_path):
    r = design_rc_beam(250, 500, 20, "C30", "HRB400",
                       M=120e6, V=180e3, l=6000)
    p = os.path.join(tmp_path, "计算书.docx")
    generate_calc_book(p, r, project="KL1")
    assert os.path.getsize(p) > 0


def test_bom_xlsx(tmp_path):
    p = os.path.join(tmp_path, "材料表.xlsx")
    generate_material_bom(p)
    assert os.path.getsize(p) > 0
    from openpyxl import load_workbook
    wb = load_workbook(p)
    assert "钢筋" in wb.sheetnames
    assert "型钢" in wb.sheetnames


def test_cli_doc_spec(tmp_path):
    from envcad import cli
    rc = cli.main(["doc", "spec", "--out", str(tmp_path), "--project", "测试"])
    assert rc == 0
    assert os.path.exists(os.path.join(tmp_path, "结构设计总说明.docx"))


def test_cli_design_rc_beam(tmp_path):
    from envcad import cli
    dxf = os.path.join(tmp_path, "beam.dxf")
    rc = cli.main(["design", "rc-beam", "--b", "250", "--h", "500",
                   "--m", "120", "--v", "180", "--l", "6000",
                   "--dxf", dxf])
    assert rc == 0
    assert os.path.exists(dxf)
