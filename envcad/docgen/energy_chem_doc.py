# -*- coding: utf-8 -*-
"""能源化工设计文档自动生成（DOCX 工艺说明书 + XLSX 设备清单）。

用法：
  from envcad.docgen.energy_chem_doc import generate_ec_doc, generate_ec_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import energy_chem_data as ec
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_ec_doc(out_path: str, project: str = "XX 能源化工项目",
                    vessel: dict = None, exchanger: dict = None,
                    tower: dict = None,
                    designer: str = "", date: str = "") -> str:
    """生成能源化工工艺设计说明书 DOCX。"""
    doc = new_cn_doc(f"{project} 工艺设计说明书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。本说明书依据现行压力容器与"
                       "工艺管道设计规范编制。")
    if designer or date:
        add_para_cn(doc, f"设计：{designer}    日期：{date}")

    add_heading_cn(doc, "二、设计依据", 1)
    rows = [[no, name] for no, name in ec.energy_chem_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if vessel:
        add_heading_cn(doc, "三、压力容器设计", 1)
        add_para_cn(doc, f"材料 {vessel['material']}，设计压力 {vessel['p']} MPa，"
                         f"内径 {vessel['Di']} mm。")
        add_table_cn(doc, ["参数", "数值"], [
            ["许用应力 [σ] (MPa)", vessel["allow"]],
            ["焊接接头系数 φ", vessel["phi"]],
            ["计算壁厚 δ (mm)", vessel["delta"]],
            ["腐蚀裕量 C2 (mm)", vessel["C2"]],
            ["名义壁厚 δn (mm)", vessel["delta_n"]],
            ["取整壁厚 (mm)", vessel["delta_n_rounded"]],
            ["计算应力 σt (MPa)", vessel["sigma_t"]],
            ["许用 [σ]φ (MPa)", vessel["allow_effective"]],
        ])
        add_para_cn(doc, vessel["note"], bold=True)

    if exchanger:
        add_heading_cn(doc, "四、换热器设计", 1)
        add_para_cn(doc, f"换热量 {exchanger['Q']} kW，"
                         f"热侧 {exchanger['hot_in']}→{exchanger['hot_out']}℃，"
                         f"冷侧 {exchanger['cold_in']}→{exchanger['cold_out']}℃。")
        add_para_cn(doc, f"对数平均温差 {exchanger['dtm']}℃，"
                         f"总传热系数 K≈{exchanger['K']} W/(m²·K)，"
                         f"估算传热面积 {exchanger['A_safety']} m²。")

    if tower:
        add_heading_cn(doc, "五、塔器设计", 1)
        add_para_cn(doc, f"填料类型 {tower['packing']}，"
                         f"泛点气速 {tower['uf']} m/s，"
                         f"操作气速 {tower['u_op']} m/s，"
                         f"塔径 {tower['D_rounded']} m。")
        add_para_cn(doc, f"Flv={tower['Flv']}，填料的泛点率={tower['ff']}，"
                         f"估算每米压降 {tower['dp_per_m']} Pa/m。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_ec_xlsx(out_path: str, project: str = "能化项目") -> str:
    """生成能源化工设备清单 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备一览表"
    ws.append(["序号", "设备名称", "规格", "材料", "数量", "备注"])
    _style(ws, 6)

    # 容器材料参考
    ws2 = wb.create_sheet("容器材料许用应力")
    ws2.append(["材料", "σb(MPa)", "σs(MPa)", "[σ](MPa)", "备注"])
    for name, p in ec.VESSEL_MATERIAL.items():
        ws2.append([name, p["sb"], p["ss"], p["allow"], p["note"]])
    _style(ws2, 5)

    # 换热K值参考
    ws3 = wb.create_sheet("换热器K值参考")
    ws3.append(["介质对", "K低(W/m²·K)", "K高(W/m²·K)"])
    for pair, (kl, kh) in ec.K_EXCHANGER_DETAIL.items():
        ws3.append([pair, kl, kh])
    _style(ws3, 3)

    wb.save(out_path)
    return out_path
