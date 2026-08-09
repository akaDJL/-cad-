# -*- coding: utf-8 -*-
"""环境应急文档自动生成（DOCX 应急预案 + XLSX 风险物质清单）。

用法：
  from envcad.docgen.emergency_doc import generate_emergency_doc
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import emergency_data as em
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_emergency_doc(out_path: str, project: str = "XX 项目",
                           risk_q: dict = None, pool: dict = None,
                           dike: dict = None, plume: dict = None,
                           designer: str = "", date: str = "") -> str:
    """生成环境应急预案说明书 DOCX。"""
    doc = new_cn_doc(f"{project} 突发环境事件应急预案")

    add_heading_cn(doc, "一、编制说明", 1)
    add_para_cn(doc, f"项目名称：{project}。本预案依据 HJ 941 等规范编制。")
    if designer or date:
        add_para_cn(doc, f"编制：{designer}    日期：{date}")

    add_heading_cn(doc, "二、编制依据", 1)
    rows = [[no, name] for no, name in em.emergency_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if risk_q:
        add_heading_cn(doc, "三、环境风险物质识别", 1)
        data = [[s["name"], s["amount"], s["q_crit"], s["qi"]]
                for s in risk_q["substances"]]
        add_table_cn(doc, ["物质", "在线量(t)", "临界量(t)", "qi=qi/Qi"], data)
        add_para_cn(doc, f"累计 Q={risk_q['Q_total']}，"
                         f"{'构成重大危险源' if risk_q['major'] else '未构成重大危险源'}。",
                    bold=True)

    if pool:
        add_heading_cn(doc, "四、事故应急池", 1)
        add_para_cn(doc, f"V1(泄漏量)={pool['V1']}m³，"
                         f"V2(消防+雨水)={pool['V2']}m³，"
                         f"V3(可转输)={pool['V3']}m³。")
        add_para_cn(doc, f"应急池有效容积 {pool['V_total']} m³。", bold=True)

    if dike:
        add_heading_cn(doc, "五、围堰/防火堤", 1)
        add_para_cn(doc, dike["note"], bold=True)

    if plume:
        add_heading_cn(doc, "六、大气风险扩散分析", 1)
        add_para_cn(doc, f"泄漏源强 {plume['Q_leak']} g/s，风速 {plume['wind_speed']} m/s，"
                         f"稳定度 {plume['stab_class']}。")
        add_para_cn(doc, f"毒性终点浓度 {plume['endpoint']} mg/m³，"
                         f"下风向风险距离约 {plume['x_max_risk']} m。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_emergency_xlsx(out_path: str, project: str = "应急项目") -> str:
    """生成环境应急风险物质清单 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "风险物质清单"
    ws.append(["序号", "物质名称", "最大存在量(t)", "临界量(t)", "Q值", "备注"])
    _style(ws, 6)

    ws2 = wb.create_sheet("风险物质临界量")
    ws2.append(["物质", "临界量(t)"])
    for name, q in em.RISK_SUBSTANCE_Q.items():
        ws2.append([name, q])
    _style(ws2, 2)

    ws3 = wb.create_sheet("毒性终点浓度")
    ws3.append(["物质", "AEGL-2 1h(mg/m³)"])
    for name, endpoint in em.TOXIC_ENDPOINT.items():
        ws3.append([name, endpoint])
    _style(ws3, 2)

    wb.save(out_path)
    return out_path
