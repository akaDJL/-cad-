# -*- coding: utf-8 -*-
"""土壤修复文档自动生成（DOCX 修复方案说明书 + XLSX 检测数据表）。

用法：
  from envcad.docgen.remediation_doc import generate_remediation_doc
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import remediation_data as rd
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_remediation_doc(out_path: str, project: str = "XX 土壤修复项目",
                             soil_results: list = None,
                             injection: dict = None,
                             thermal: dict = None,
                             designer: str = "", date: str = "") -> str:
    """生成土壤修复方案说明书 DOCX。

    soil_results: [dict(pollutant, concentration, land_type, ok, ratio, note), ...]
    """
    doc = new_cn_doc(f"{project} 土壤修复方案说明书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。本方案依据 HJ 25 系列导则"
                       "与 GB 36600 建设用地土壤风险管控标准编制。")
    if designer or date:
        add_para_cn(doc, f"编制：{designer}    日期：{date}")

    add_heading_cn(doc, "二、编制依据", 1)
    rows = [[no, name] for no, name in rd.remediation_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if soil_results:
        add_heading_cn(doc, "三、污染现状与达标判定", 1)
        header = ["污染物", "实测(mg/kg)", "筛选值(mg/kg)", "超标倍数", "达标"]
        data = [[r["pollutant"], r["concentration"], r["limit"],
                 f'{r["ratio"]}×' if not r["ok"] else "—",
                 "√" if r["ok"] else "×"]
                for r in soil_results if r["limit"] != "未知"]
        add_table_cn(doc, header, data)
        for r in soil_results:
            if r["ok"] is False:
                add_para_cn(doc, r["note"], bold=True)

    if injection:
        add_heading_cn(doc, "四、原位注入方案", 1)
        add_para_cn(doc, f"渗透系数 k={injection['k']} m/d，"
                         f"孔隙度 n={injection['porosity']}。")
        add_para_cn(doc, f"注入流量 {injection['inject_rate']} L/min，"
                         f"影响半径约 {injection['R_m']} m，"
                         f"建议井间距 {injection['spacing_m']} m。")

    if thermal:
        add_heading_cn(doc, "五、热脱附参数", 1)
        add_para_cn(doc, f"污染物类型 {thermal['pollutant_type']}，"
                         f"推荐 {thermal['level']}。")
        add_para_cn(doc, f"操作温度 {thermal['operating_temp']}℃，"
                         f"停留时间 {thermal['dwell_time']} min。")
        add_para_cn(doc, thermal["note"], bold=True)

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_remediation_xlsx(out_path: str, project: str = "修复项目") -> str:
    """生成土壤检测数据 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "土壤检测结果"
    ws.append(["序号", "污染物", "实测(mg/kg)", "第一类筛选值", "第二类筛选值", "备注"])
    _style(ws, 6)

    ws2 = wb.create_sheet("筛选值参考")
    ws2.append(["污染物", "第一类(mg/kg)", "第二类(mg/kg)", "类别"])
    for name, p in rd.SOIL_SCREENING.items():
        ws2.append([name, p.get("I", ""), p.get("II", ""), p.get("note", "")])
    _style(ws2, 4)

    ws3 = wb.create_sheet("热脱附参数")
    ws3.append(["等级", "温度范围(℃)", "停留时间(min)", "适用"]
               )
    for level, p in rd.THERMAL_DESORPTION.items():
        ws3.append([level, f'{p["temp"][0]}~{p["temp"][1]}',
                    f'{p["dwell"][0]}~{p["dwell"][1]}', p["note"]])
    _style(ws3, 4)

    wb.save(out_path)
    return out_path
