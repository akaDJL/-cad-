# -*- coding: utf-8 -*-
"""农业食品机械设计文档自动生成（DOCX 说明书 + XLSX 设备表）。

用法：
  from envcad.docgen.agri_doc import generate_agri_doc, generate_agri_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import agri_data as ag
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_agri_doc(out_path: str, project: str = "XX 农业装备项目",
                      irrigation: dict = None, screw: dict = None,
                      packaging: dict = None,
                      designer: str = "", date: str = "") -> str:
    """生成农食装备设计说明书 DOCX。"""
    doc = new_cn_doc(f"{project} 农食装备设计说明书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。")
    if designer or date:
        add_para_cn(doc, f"设计：{designer}    日期：{date}")

    add_heading_cn(doc, "二、设计依据", 1)
    rows = [[no, name] for no, name in sorted(ag.AGRI_CODES.items())]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if irrigation:
        add_heading_cn(doc, "三、灌溉系统", 1)
        add_para_cn(doc, f"灌溉面积 {irrigation['area_mu']} 亩，作物 {irrigation['crop']}，"
                         f"{irrigation['method']}方式。")
        add_para_cn(doc, f"日需水量 {irrigation['Q_day']} m³/d，"
                         f"系统设计流量 {irrigation['Q_sys']} m³/h，"
                         f"推荐管径 DN{irrigation['d_select']}。")

    if screw:
        add_heading_cn(doc, "四、螺旋输送机", 1)
        add_para_cn(doc, f"输送量 {screw['Q']} t/h，长度 {screw['L']} m，"
                         f"物料 {screw['material_type']}。")
        add_para_cn(doc, f"螺旋直径 D{screw['D_selected']} mm，"
                         f"轴功率 {screw['P_axis']} kW。")

    if packaging:
        add_heading_cn(doc, "五、包装系统", 1)
        add_para_cn(doc, f"包装机类型 {packaging['machine_type']}，"
                         f"袋长 {packaging['bag_length']} mm。")
        add_para_cn(doc, f"理论产能 {packaging['speed_range']}。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_agri_xlsx(out_path: str, project: str = "农食装备") -> str:
    """生成农食装备设备清单 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备一览表"
    ws.append(["序号", "名称", "规格", "数量", "备注"])
    _style(ws, 5)

    ws2 = wb.create_sheet("灌溉参数参考")
    ws2.append(["作物", "日耗水强度(mm/d)"])
    for crop, et in ag.IRRIGATION_ET.items():
        ws2.append([crop, et])
    _style(ws2, 2)

    ws3 = wb.create_sheet("螺旋输送机参数")
    ws3.append(["物料", "填充系数ψ", "阻力系数ω"])
    for m, v in ag.SCREW_FILL.items():
        ws3.append([m, v, "-"])
    _style(ws3, 3)

    wb.save(out_path)
    return out_path
