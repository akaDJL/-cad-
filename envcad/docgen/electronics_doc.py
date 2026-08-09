# -*- coding: utf-8 -*-
"""电子硬件设计文档自动生成（DOCX 设计说明书 + XLSX BOM 表）。

用法：
  from envcad.docgen.electronics_doc import generate_elec_doc, generate_elec_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import electronics_data as ed
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_elec_doc(out_path: str, project: str = "XX 电子设备",
                      pcb: dict = None, thermal: dict = None,
                      impedance: dict = None,
                      designer: str = "", date: str = "") -> str:
    """生成电子硬件设计说明书 DOCX。"""
    doc = new_cn_doc(f"{project} 硬件设计说明书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。本说明书依据 IPC 标准与"
                       "JEDEC 封装规范编制。")
    if designer or date:
        add_para_cn(doc, f"设计：{designer}    日期：{date}")

    add_heading_cn(doc, "二、设计依据", 1)
    rows = [[no, name] for no, name in sorted(ed.ELECTRONICS_CODES.items())]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if pcb:
        add_heading_cn(doc, "三、PCB 走线载流", 1)
        add_para_cn(doc, f"走线宽度 {pcb['width']}mm，铜厚 {pcb['copper_oz']}，"
                         f"{pcb['layer']}层，温升 {pcb['delta_T']}℃。")
        add_para_cn(doc, f"估算载流量 {pcb['I']} A。")
        add_table_cn(doc, ["参数", "数值"], [
            ["走线宽度 (mm)", pcb["width"]],
            ["铜厚", pcb["copper_oz"]],
            ["铜厚 (μm)", pcb["thickness_um"]],
            ["层别", pcb["layer"]],
            ["允许温升 (℃)", pcb["delta_T"]],
            ["估算载流量 (A)", pcb["I"]],
        ])

    if thermal:
        add_heading_cn(doc, "四、散热设计", 1)
        add_para_cn(doc, f"耗散功率 {thermal['Pd']}W，结温限制 {thermal['T_j_max']}℃，"
                         f"环境温度 {thermal['T_amb']}℃。")
        add_para_cn(doc, f"散热器允许热阻 {thermal['R_hs']}℃/W，"
                         f"所需散热面积约 {thermal['A_cm2']} cm²。")
        add_para_cn(doc, thermal["note"], bold=True)

    if impedance:
        add_heading_cn(doc, "五、微带线阻抗", 1)
        add_para_cn(doc, f"介质 εr={impedance['er']}，厚度 {impedance['h']}mm，"
                         f"走线宽 {impedance['w']}mm。")
        add_para_cn(doc, f"特性阻抗 Z0≈{impedance['Z0']}Ω。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_elec_xlsx(out_path: str, project: str = "电子设备") -> str:
    """生成电子设备 BOM XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "元器件清单"
    ws.append(["序号", "名称", "型号/规格", "封装", "数量", "备注"])
    _style(ws, 6)

    ws2 = wb.create_sheet("IC封装参考")
    ws2.append(["封装", "本体(mm)", "引脚间距(mm)", "引脚数", "备注"])
    for name, p in ed.IC_PACKAGE.items():
        ws2.append([name, f'{p["body"][0]}×{p["body"][1]}',
                    p["pitch"], p["count"], p["note"]])
    _style(ws2, 5)

    ws3 = wb.create_sheet("连接器规格")
    ws3.append(["系列", "间距(mm)", "排数", "引脚范围", "备注"])
    for name, p in ed.CONNECTOR.items():
        ws3.append([name, p["pitch"], p["row"],
                    f'{p["pin_count"][0]}~{p["pin_count"][1]}', p["note"]])
    _style(ws3, 5)

    ws4 = wb.create_sheet("IP防护等级")
    ws4.append(["等级", "防尘", "防水", "适用场景"])
    for code, p in ed.IP_CODE.items():
        ws4.append([code, p["dust"], p["water"], p["use"]])
    _style(ws4, 4)

    wb.save(out_path)
    return out_path
