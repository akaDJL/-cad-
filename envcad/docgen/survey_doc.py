# -*- coding: utf-8 -*-
"""测绘GIS文档自动生成（DOCX 技术设计书 + XLSX 控制点成果表）。

用法：
  from envcad.docgen.survey_doc import generate_survey_doc, generate_survey_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import survey_data as sv
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_survey_doc(out_path: str, project: str = "XX 测绘项目",
                        gauss: dict = None, accuracy: dict = None,
                        pipe_check: dict = None,
                        designer: str = "", date: str = "") -> str:
    """生成测绘技术设计书 DOCX。"""
    doc = new_cn_doc(f"{project} 测绘技术设计书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。采用 CGCS2000 国家大地坐标系，"
                       "高斯-克吕格投影。")
    if designer or date:
        add_para_cn(doc, f"编制：{designer}    日期：{date}")

    add_heading_cn(doc, "二、技术依据", 1)
    rows = [[no, name] for no, name in sorted(sv.SURVEY_CODES.items())]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if gauss:
        add_heading_cn(doc, "三、坐标系统与投影", 1)
        add_para_cn(doc, f"椭球：{sv.CGCS2000['name']}，"
                         f"长半轴 a={sv.CGCS2000['a']}m，扁率 f={sv.CGCS2000['f']:.10f}。")
        add_para_cn(doc, f"高斯投影 {gauss['zone']}°带，"
                         f"中央子午线 L0={gauss['L0']}°，"
                         f"x={gauss['x']}m，y={gauss['y_nat']}m。")

    if accuracy:
        add_heading_cn(doc, "四、精度评定", 1)
        add_para_cn(doc, f"{accuracy['grade']}控制：平面允许 {accuracy['standard_h']}mm，"
                         f"实测 {accuracy['measured_h']}mm，"
                         f"{'√' if accuracy['h_ok'] else '×'}。")
        add_para_cn(doc, f"高程允许 {accuracy['standard_v']}mm，"
                         f"实测 {accuracy['measured_v']}mm，"
                         f"{'√' if accuracy['v_ok'] else '×'}。")

    if pipe_check:
        add_heading_cn(doc, "五、管线探测", 1)
        add_para_cn(doc, f"{pipe_check['pipe_type']}管线，最小埋深 {pipe_check['min_depth']}m，"
                         f"实际埋深 {pipe_check['actual_depth']}m，"
                         f"{'满足' if pipe_check['ok'] else '不满足'}要求。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_survey_xlsx(out_path: str, project: str = "测绘项目") -> str:
    """生成测绘控制点成果表 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "控制点成果表"
    ws.append(["点名", "等级", "X(m)", "Y(m)", "H(m)", "备注"])
    _style(ws, 6)

    ws2 = wb.create_sheet("精度等级参考")
    ws2.append(["等级", "平面中误差(mm)", "高程中误差(mm)", "用途"])
    for grade, p in sv.CONTROL_ACCURACY.items():
        ws2.append([grade, p["h"], p["v"], p["use"]])
    _style(ws2, 4)

    ws3 = wb.create_sheet("比例尺精度")
    ws3.append(["比例尺", "地物间距(m)", "等高距(m)", "用途"])
    for scale, p in sv.MAP_SCALES.items():
        ws3.append([scale, p["max_spacing"], p["contour"], p["use"]])
    _style(ws3, 4)

    wb.save(out_path)
    return out_path
