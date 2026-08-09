# -*- coding: utf-8 -*-
"""电气设计文档自动生成（DOCX 设计说明书 + XLSX 负荷计算表）。

用法：
  from envcad.docgen.elec_doc import generate_elec_spec, generate_load_xlsx
  generate_elec_spec("电气设计说明.docx", project="XX配电工程",
                     load=load, cable=cable, illum=illum, sc=sc)
  generate_load_xlsx("负荷计算表.xlsx", items=[...])
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import elec_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_elec_spec(out_path: str, project: str = "XX 电气工程",
                       load: dict = None, cable: dict = None,
                       illum: dict = None, sc: dict = None) -> str:
    """生成电气设计说明书。load/cable/illum/sc 为 design.electrical 结果。"""
    doc = new_cn_doc(f"{project} 电气设计说明")

    add_heading_cn(doc, "一、工程概况", 1)
    add_para_cn(doc, f"工程名称：{project}")
    add_para_cn(doc, "本工程为低压配电系统设计，供电电压 380/220V，"
                     "中性点直接接地（TN-S 系统），负荷等级按二级考虑。")

    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本工程依据下列现行国家标准、规范：")
    add_table_cn(doc, ["标准编号", "名称"],
                 [[no, name] for no, name in elec_data.elec_code_list()])

    add_heading_cn(doc, "三、负荷计算", 1)
    if load:
        add_para_cn(doc, f"采用需要系数法计算：用电性质「{load['kind']}」，"
                         f"安装容量 Pe={load['Pe']}kW，需要系数 Kx={load['Kx']}，"
                         f"功率因数 cosφ={load['cos']}。")
        add_table_cn(doc, ["参数", "数值"], [
            ["计算有功 Pjs", f"{load['Pjs']} kW"],
            ["计算无功 Qjs", f"{load['Qjs']} kvar"],
            ["视在功率 Sjs", f"{load['Sjs']} kVA"],
            ["计算电流 Ijs", f"{load['Ijs']} A"],
        ])

    add_heading_cn(doc, "四、导线电缆选择", 1)
    if cable:
        add_para_cn(doc, cable["note"])
        add_para_cn(doc, "结论：" + ("所选电缆满足载流量与允许电压降要求。"
                                     if cable["ok"] else "需加大截面或缩短线路。"))

    add_heading_cn(doc, "五、照明设计", 1)
    if illum:
        add_para_cn(doc, illum["note"])

    add_heading_cn(doc, "六、短路电流与保护", 1)
    if sc:
        add_para_cn(doc, sc["note"])
    add_para_cn(doc, "低压配电设置断路器作短路与过载保护，设备外露可导电部分"
                     "可靠接地（PE），并按规范设置剩余电流保护（RCD）。")

    add_heading_cn(doc, "七、防雷与接地", 1)
    add_para_cn(doc, "建筑物按第三类防雷设防，采用避雷带/网+引下线+接地装置；"
                     "工作接地、保护接地、防雷接地共用接地体，接地电阻 ≤4Ω。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_load_xlsx(out_path: str, items: list = None) -> str:
    """生成负荷计算表。items 为 [{name, Pe, kind}] 列表；缺省给示例。"""
    from ..design.electrical import design_power_load
    wb = Workbook()
    ws = wb.active
    ws.title = "负荷计算表"
    ws.append(["序号", "回路/设备", "装机容量Pe(kW)", "用电性质",
               "需要系数Kx", "cosφ", "计算有功(kW)", "视在(kVA)", "计算电流(A)"])
    items = items or [
        dict(name="照明插座", Pe=45, kind="办公照明"),
        dict(name="空调动力", Pe=120, kind="空调"),
        dict(name="电梯", Pe=30, kind="电梯"),
        dict(name="消防水泵", Pe=55, kind="水泵"),
    ]
    tot_p = tot_s = 0.0
    for i, it in enumerate(items, 1):
        try:
            r = design_power_load(it["Pe"], it.get("kind", "办公照明"))
        except Exception as _e:
            continue
        ws.append([i, it["name"], r["Pe"], r["kind"], r["Kx"], r["cos"],
                   r["Pjs"], r["Sjs"], r["Ijs"]])
        tot_p += r["Pjs"]
        tot_s += r["Sjs"]
    ws.append(["", "合计", "", "", "", "", round(tot_p, 1), round(tot_s, 1), ""])
    _style(ws, 9)
    for col, w in zip("ABCDEFGHI", [6, 14, 15, 12, 11, 8, 12, 10, 12]):
        ws.column_dimensions[col].width = w
    wb.save(out_path)
    return out_path
