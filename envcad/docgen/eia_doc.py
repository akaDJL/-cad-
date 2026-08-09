# -*- coding: utf-8 -*-
"""环境影响评价文档自动生成（DOCX 环评报告表 + XLSX 评价等级汇总表）。

用法：
  from envcad.docgen.eia_doc import generate_eia_doc, generate_eia_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import eia_data as ea
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_eia_doc(out_path: str, project: str = "XX 建设项目",
                     air_grade: dict = None, water_grade: dict = None,
                     noise_grade: dict = None, prot_dist: dict = None,
                     sensitive: dict = None,
                     designer: str = "", date: str = "") -> str:
    """生成环评报告表（简化版） DOCX。"""
    doc = new_cn_doc(f"{project} 环境影响评价报告表")

    add_heading_cn(doc, "一、项目概况", 1)
    add_para_cn(doc, f"项目名称：{project}。本报告依据 HJ 2.1 等环评技术导则编制。")
    if designer or date:
        add_para_cn(doc, f"编制：{designer}    日期：{date}")

    add_heading_cn(doc, "二、编制依据", 1)
    rows = [[no, name] for no, name in ea.eia_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    add_heading_cn(doc, "三、评价等级判定", 1)

    if air_grade:
        add_heading_cn(doc, "3.1 大气环境", 2)
        add_para_cn(doc, f"Pmax={air_grade['pmax']}% D10%={air_grade['d10']}m → "
                         f"大气评价{air_grade['grade']}。")
        add_para_cn(doc, f"预测方法：{air_grade['method']}。")

    if water_grade:
        add_heading_cn(doc, "3.2 地表水环境", 2)
        add_para_cn(doc, water_grade["note"])

    if noise_grade:
        add_heading_cn(doc, "3.3 声环境", 2)
        add_para_cn(doc, noise_grade["note"])

    # 汇总表
    rows = []
    for g in [air_grade, water_grade, noise_grade]:
        if g:
            rows.append([g.get("note", "").split("→")[0].strip(),
                         g["grade"],
                         g.get("method", g.get("detail", ""))])
    if rows:
        add_heading_cn(doc, "四、评价等级汇总", 1)
        add_table_cn(doc, ["环境要素", "评价等级", "方法/要求"], rows)

    if prot_dist:
        add_heading_cn(doc, "五、防护距离", 1)
        add_para_cn(doc, f"计算防护距离 {prot_dist['d_calc']} m → "
                         f"取整 {prot_dist['d_protection']} m。")

    if sensitive:
        add_heading_cn(doc, "六、敏感区距离", 1)
        add_para_cn(doc, f"{sensitive['area_name']}({sensitive['level']}类)，"
                         f"缓冲距离 {sensitive['buffer_min']}m，"
                         f"实际 {sensitive['actual_distance']}m，"
                         f"{'满足' if sensitive['ok'] else '不满足'}。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_eia_xlsx(out_path: str, project: str = "环评项目") -> str:
    """生成环评等级汇总 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "评价等级判定"
    ws.append(["环境要素", "Pmax/排放量/增量", "评价等级", "方法"])
    _style(ws, 4)

    ws2 = wb.create_sheet("评价等级标准(大气)")
    ws2.append(["等级", "Pmax", "D10%", "详情"])
    for grade, info in ea.AIR_GRADE.items():
        ws2.append([grade, info["Pmax"], info["D10%"], info["detail"]])
    _style(ws2, 4)

    ws3 = wb.create_sheet("环境敏感区")
    ws3.append(["名称", "级别", "缓冲距离(m)"])
    for name, p in ea.SENSITIVE_AREA.items():
        ws3.append([name, p["level"],
                    f'{p["buffer"][0]}~{p["buffer"][1]}'])
    _style(ws3, 3)

    ws4 = wb.create_sheet("总量控制指标")
    ws4.append(["指标", "含义"])
    for k, v in ea.TOTAL_CONTROL.items():
        ws4.append([k, v])
    _style(ws4, 2)

    wb.save(out_path)
    return out_path
