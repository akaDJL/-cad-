# -*- coding: utf-8 -*-
"""暖通空调设计文档自动生成（DOCX 设计说明书 + XLSX 负荷设备表）。

用法：
  from envcad.docgen.hvac_doc import generate_hvac_spec, generate_hvac_xlsx
  generate_hvac_spec("暖通设计说明.docx", project="XX空调工程",
                     load=load, air=air, fresh=fresh, duct=duct)
  generate_hvac_xlsx("负荷设备表.xlsx", items=[...])
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import hvac_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_hvac_spec(out_path: str, project: str = "XX 暖通工程",
                       load: dict = None, air: dict = None,
                       fresh: dict = None, duct: dict = None) -> str:
    """生成暖通空调设计说明书。各参数为 design.hvac 结果。"""
    doc = new_cn_doc(f"{project} 暖通空调设计说明")

    add_heading_cn(doc, "一、工程概况", 1)
    add_para_cn(doc, f"工程名称：{project}")
    add_para_cn(doc, "本工程设舒适性空调与通风系统，冷源采用冷水机组，"
                     "热源接入市政热力，末端为风机盘管加新风。")

    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本工程依据下列现行国家标准、规范：")
    add_table_cn(doc, ["标准编号", "名称"],
                 [[no, name] for no, name in hvac_data.hvac_code_list()])

    add_heading_cn(doc, "三、冷热负荷", 1)
    if load:
        add_para_cn(doc, load["note"])
        add_table_cn(doc, ["项目", "指标", "负荷"], [
            ["冷负荷", f"{load['q_cool']} W/m²", f"{load['Qc']} kW"],
            ["热负荷", f"{load['q_heat']} W/m²", f"{load['Qh']} kW"],
        ])

    add_heading_cn(doc, "四、送风量与新风", 1)
    if air:
        add_para_cn(doc, "送风量：" + air["note"])
    if fresh:
        add_para_cn(doc, "新风量：" + fresh["note"])

    add_heading_cn(doc, "五、风管系统", 1)
    if duct:
        add_para_cn(doc, "风管尺寸：" + duct["note"])
    add_para_cn(doc, "风管采用镀锌钢板制作，按规范做保温；风系统设置防火阀，"
                     "穿越防火分区处设 70℃ 熔断防火阀。")

    add_heading_cn(doc, "六、节能与防排烟", 1)
    add_para_cn(doc, "空调水系统采用变流量运行；地下车库、内走道等按规范设置"
                     "机械排烟系统，排烟量按现行《建筑防烟排烟系统技术标准》GB 51251 计算。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_hvac_xlsx(out_path: str, items: list = None) -> str:
    """生成分区负荷/设备表。items 为 [{name, area, place, height}] 列表。"""
    from ..design.hvac import design_load, design_air_volume
    wb = Workbook()
    ws = wb.active
    ws.title = "分区负荷设备表"
    ws.append(["序号", "房间/分区", "面积(m²)", "功能", "冷负荷(kW)",
               "热负荷(kW)", "换气次数", "送风量(m³/h)"])
    items = items or [
        dict(name="办公区", area=800, place="办公室", height=3.0),
        dict(name="会议室", area=200, place="会议室", height=3.2),
        dict(name="大堂", area=300, place="商场", height=4.5),
    ]
    tot_c = tot_h = 0.0
    for i, it in enumerate(items, 1):
        try:
            r = design_load(it["area"], it.get("place", "办公室"))
            a = design_air_volume(it["area"], it.get("height", 3.0),
                                  it.get("place", "办公室"))
        except Exception as _e:
            continue
        ws.append([i, it["name"], r["area"], r["place"], r["Qc"],
                   r["Qh"], a["n"], a["L"]])
        tot_c += r["Qc"]
        tot_h += r["Qh"]
    ws.append(["", "合计", "", "", round(tot_c, 1), round(tot_h, 1), "", ""])
    _style(ws, 8)
    for col, w in zip("ABCDEFGH", [6, 14, 10, 10, 12, 12, 10, 13]):
        ws.column_dimensions[col].width = w
    wb.save(out_path)
    return out_path
