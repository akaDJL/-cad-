# -*- coding: utf-8 -*-
"""环保工艺设计文档自动生成（DOCX 工艺说明书 + XLSX 排放达标清单）。

用法：
  from envcad.docgen.env_report import generate_env_spec, generate_discharge_xlsx
  generate_env_spec("工艺说明.docx", project="XX污水处理厂",
                    aeration=aer, sed=sed, dust=dust)
  generate_discharge_xlsx("排放清单.xlsx", standard="一级A")
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import env_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_env_spec(out_path: str, project: str = "XX 环保工程",
                      discharge_std: str = "一级A",
                      noise_zone: str = "3类",
                      aeration: dict = None, sed: dict = None,
                      dust: dict = None) -> str:
    """生成环保工艺设计说明书。aeration/sed/dust 为 design.env_process 结果。"""
    doc = new_cn_doc(f"{project} 环保工艺设计说明")

    # 1 工程概况
    add_heading_cn(doc, "一、工程概况", 1)
    add_para_cn(doc, f"工程名称：{project}")
    add_para_cn(doc, f"排放执行标准：{discharge_std}；厂界噪声执行 {noise_zone}声环境功能区。")

    # 2 设计依据
    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本工程遵照下列现行环境保护标准、规范：")
    rows = [[no, name] for no, name in env_data.env_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    # 3 设计进出水水质
    add_heading_cn(doc, "三、排放限值要求", 1)
    add_para_cn(doc, f"出水执行 GB 18918《城镇污水处理厂污染物排放标准》{discharge_std} 标准：")
    lim = env_data.water_limit(discharge_std)
    rows = [[k, v] for k, v in lim.items()]
    add_table_cn(doc, ["指标", "限值(mg/L, pH除外)"], rows)

    # 4 主要构筑物工艺设计
    add_heading_cn(doc, "四、主要构筑物工艺设计", 1)
    if aeration:
        add_heading_cn(doc, "4.1 曝气池", 2)
        add_para_cn(doc, f"设计流量 Q={aeration['Q']} m³/d，进水 BOD5={aeration['So']} mg/L，"
                         f"出水 BOD5={aeration['Se']} mg/L。")
        add_para_cn(doc, f"采用污泥负荷法：污泥负荷 {aeration['Ls']} kgBOD/(kgMLSS·d)，"
                         f"MLSS={aeration['MLSS']} mg/L。")
        add_para_cn(doc, f"计算容积 V={aeration['V']} m³，水力停留时间 HRT={aeration['HRT']} h，"
                         f"BOD 去除率 {aeration['removal']}%。")
    if sed:
        add_heading_cn(doc, "4.2 二次沉淀池", 2)
        add_para_cn(doc, f"表面负荷 {sed['q']} m³/(m²·h)，设计 {sed['n']} 座，"
                         f"每座直径 Φ{sed['D']} m，有效水深 {sed['depth']} m，"
                         f"总沉淀面积 {sed['A_total']} m²。")
    if dust:
        add_heading_cn(doc, "4.3 除尘系统", 2)
        add_para_cn(doc, dust["note"])

    # 5 污染物排放与环境影响
    add_heading_cn(doc, "五、噪声控制", 1)
    nz = env_data.noise_limit(noise_zone)
    add_para_cn(doc, f"厂界噪声按 {noise_zone}（{nz['use']}）控制：昼间 ≤{nz['day']} dB(A)，"
                     f"夜间 ≤{nz['night']} dB(A)。风机、水泵等设备采取隔声减振措施。")

    # 6 运行管理要求
    add_heading_cn(doc, "六、运行管理要求", 1)
    add_para_cn(doc, "1）在线监测 COD、氨氮、pH、流量并联网上传；"
                     "2）污泥定期外运至有资质单位处置，禁止随意堆放；"
                     "3）事故池容积满足非正常工况调蓄；"
                     "4）各排放口规范化设置并标识。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_discharge_xlsx(out_path: str, standard: str = "一级A",
                            actual: dict = None) -> str:
    """生成污染物排放达标对照清单。actual 为各指标实测/预测浓度 dict。"""
    wb = Workbook()

    # 水污染物
    ws1 = wb.active
    ws1.title = "水污染物"
    ws1.append(["序号", "污染物", f"限值({standard})", "设计/预测值", "达标"])
    lim = env_data.water_limit(standard)
    actual = actual or {}
    i = 1
    for k, v in lim.items():
        av = actual.get(k, "")
        ok = ""
        if isinstance(v, (int, float)) and isinstance(av, (int, float)):
            ok = "达标" if av <= v else "超标"
        ws1.append([i, k, v, av, ok])
        i += 1
    _style(ws1, 5)
    for col, w in zip("ABCDE", [6, 14, 16, 14, 10]):
        ws1.column_dimensions[col].width = w

    # 大气污染物
    ws2 = wb.create_sheet("大气污染物")
    ws2.append(["序号", "污染物", "限值(mg/m³)", "说明"])
    for i, (k, p) in enumerate(env_data.AIR_GB16297.items(), 1):
        ws2.append([i, k, p["conc"], p["note"]])
    _style(ws2, 4)
    for col, w in zip("ABCD", [6, 16, 16, 20]):
        ws2.column_dimensions[col].width = w

    # 噪声
    ws3 = wb.create_sheet("厂界噪声")
    ws3.append(["功能区", "昼间 dB(A)", "夜间 dB(A)", "适用区域"])
    for k, p in env_data.NOISE_GB3096.items():
        ws3.append([k, p["day"], p["night"], p["use"]])
    _style(ws3, 4)
    for col, w in zip("ABCD", [10, 14, 14, 18]):
        ws3.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
