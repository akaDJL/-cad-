# -*- coding: utf-8 -*-
"""机械设计文档自动生成（DOCX 计算说明书 + XLSX 零件明细表）。

用法：
  from envcad.docgen.mech_calc import generate_mech_calc, generate_parts_xlsx
  generate_mech_calc("机械设计计算说明书.docx", project="XX减速器",
                     gear=gear_result, shaft=shaft_result)
  generate_parts_xlsx("零件明细表.xlsx", parts=[...])

gear/shaft 分别为 design.gear.check_spur_gear 与 design.shaft.design_shaft
的返回结果 dict。未提供时相应章节自动跳过。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import mech_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_mech_calc(out_path: str, project: str = "XX 机械传动装置",
                       gear: dict = None, shaft: dict = None,
                       designer: str = "", date: str = "") -> str:
    """生成机械设计计算说明书 DOCX。

    参数：
        gear   design.gear.check_spur_gear() 结果
        shaft  design.shaft.design_shaft() 结果
    """
    doc = new_cn_doc(f"{project} 机械设计计算说明书")

    # 一、设计概述
    add_heading_cn(doc, "一、设计概述", 1)
    add_para_cn(doc, f"设计对象：{project}。")
    add_para_cn(doc, "本说明书依据机械设计相关国家标准，对传动关键零件进行"
                     "强度校核与选型计算，作为设计与制造依据。")
    if designer or date:
        add_para_cn(doc, f"设计：{designer}    日期：{date}")

    # 二、设计依据
    add_heading_cn(doc, "二、设计依据（现行标准）", 1)
    rows = [[no, name] for no, name in mech_data.mech_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    # 三、齿轮传动强度校核
    if gear:
        add_heading_cn(doc, "三、齿轮传动强度校核", 1)
        add_para_cn(doc, f"小齿轮材料 {gear['material']}，传递功率 P={gear['power']} kW，"
                         f"转速 n1={gear['n1']} rpm。")
        add_heading_cn(doc, "3.1 几何参数", 2)
        add_table_cn(doc, ["参数", "数值"], [
            ["齿数 z1 / z2", f"{gear['z1']} / {gear['z2']}"],
            ["传动比 u", gear["u"]],
            ["模数 m (mm)", gear["mn"]],
            ["分度圆 d1 / d2 (mm)", f"{gear['d1']} / {gear['d2']}"],
            ["中心距 a (mm)", gear["a"]],
            ["齿宽 b (mm)", gear["b"]],
            ["小齿轮扭矩 T1 (N·mm)", f"{gear['T1']:.0f}"],
        ])
        add_heading_cn(doc, "3.2 接触疲劳强度", 2)
        add_para_cn(doc, f"计算接触应力 σH={gear['sH']} N/mm²，许用 [σH]={gear['sHP']} N/mm²，"
                         f"安全系数 {gear['SH_calc']}，"
                         f"{'满足' if gear['sH_ok'] else '不满足'}。")
        add_heading_cn(doc, "3.3 齿根弯曲疲劳强度", 2)
        add_para_cn(doc, f"计算弯曲应力 σF={gear['sF']} N/mm²，许用 [σF]={gear['sFP']} N/mm²，"
                         f"安全系数 {gear['SF_calc']}，"
                         f"{'满足' if gear['sF_ok'] else '不满足'}。")
        add_para_cn(doc, "结论：" + ("齿轮接触与弯曲强度均满足要求。"
                    if gear["all_ok"] else "齿轮强度不足，应增大模数/齿宽或提高材料。"),
                    bold=True)

    # 四、轴的强度设计
    if shaft:
        add_heading_cn(doc, "四、轴的强度设计与校核", 1)
        est = shaft["estimate"]
        chk = shaft["check"]
        add_heading_cn(doc, "4.1 按扭转强度初估轴径", 2)
        add_para_cn(doc, f"材料 {est['material']}，系数 A0={est['A0']}，"
                         f"P={est['power']} kW，n={est['n']} rpm。")
        add_para_cn(doc, f"初估最小轴径 d≥{est['d_calc']} mm，圆整取标准直径 d={est['d']} mm。")
        add_heading_cn(doc, "4.2 弯扭合成强度校核", 2)
        add_table_cn(doc, ["参数", "数值"], [
            ["校核直径 d (mm)", chk["d"]],
            ["合成弯矩 M (N·mm)", f"{chk['M']:.0f}"],
            ["扭矩 T (N·mm)", f"{chk['T']:.0f}"],
            ["折合系数 α", chk["alpha"]],
            ["抗弯截面系数 W (mm³)", chk["W"]],
            ["计算应力 σca (N/mm²)", chk["sigma_ca"]],
            ["许用应力 [σ-1b] (N/mm²)", chk["allow"]],
            ["安全系数", chk["safety"]],
        ])
        add_para_cn(doc, "结论：" + ("轴径满足弯扭合成强度要求。"
                    if shaft["all_ok"] else "轴径不足，应增大直径或提高材料。"),
                    bold=True)

    # 五、结论
    add_heading_cn(doc, "五、总结论", 1)
    parts_ok = []
    if gear:
        parts_ok.append("齿轮传动" + ("满足" if gear["all_ok"] else "不满足"))
    if shaft:
        parts_ok.append("传动轴" + ("满足" if shaft["all_ok"] else "不满足"))
    if parts_ok:
        add_para_cn(doc, "、".join(parts_ok) + "强度要求。")
    else:
        add_para_cn(doc, "本次未提供具体校核数据。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


# 缺省零件明细示例（可被 parts 参数覆盖）
_DEMO_PARTS = [
    dict(no=1, name="小齿轮", qty=1, material="40Cr", spec="m=2 z=20", remark="调质"),
    dict(no=2, name="大齿轮", qty=1, material="45钢", spec="m=2 z=60", remark="调质"),
    dict(no=3, name="传动轴", qty=1, material="45钢", spec="Φ35", remark="调质"),
    dict(no=4, name="平键", qty=2, material="45钢", spec="GB/T 1096", remark="10×8"),
    dict(no=5, name="深沟球轴承", qty=2, material="轴承钢", spec="6207", remark="GB/T 276"),
]


def generate_parts_xlsx(out_path: str, parts: list = None,
                        project: str = "机械装置") -> str:
    """生成零件明细表 XLSX。

    parts: [dict(no, name, qty, material, spec, remark), ...]
    另附标准数据参考页（材料许用应力 / 标准模数 / 螺纹）。
    """
    wb = Workbook()

    # Sheet1 零件明细
    ws = wb.active
    ws.title = "零件明细表"
    ws.append(["序号", "名称", "数量", "材料", "规格/型号", "备注"])
    parts = parts or _DEMO_PARTS
    for p in parts:
        ws.append([p.get("no", ""), p.get("name", ""), p.get("qty", ""),
                   p.get("material", ""), p.get("spec", ""), p.get("remark", "")])
    _style(ws, 6)
    for col, w in zip("ABCDEF", [6, 16, 8, 14, 18, 16]):
        ws.column_dimensions[col].width = w

    # Sheet2 材料力学性能参考
    ws2 = wb.create_sheet("材料力学性能")
    ws2.append(["材料", "σb(N/mm²)", "σs(N/mm²)", "硬度HB", "接触疲劳σHlim",
                "弯曲疲劳σFlim", "备注"])
    for name, p in mech_data.MECH_MATERIAL.items():
        ws2.append([name, p["sb"], p["ss"], p["HB"], p["sH_lim"],
                    p["sF_lim"], p["note"]])
    _style(ws2, 7)
    for col, w in zip("ABCDEFG", [16, 12, 12, 10, 14, 14, 14]):
        ws2.column_dimensions[col].width = w

    # Sheet3 标准模数与直径
    ws3 = wb.create_sheet("标准模数与直径")
    ws3.append(["标准模数第一系列(mm)", ", ".join(str(m) for m in mech_data.MODULE_1)])
    ws3.append(["标准模数第二系列(mm)", ", ".join(str(m) for m in mech_data.MODULE_2)])
    ws3.append(["标准直径系列(mm)", ", ".join(str(d) for d in mech_data.STD_DIAMETER)])
    ws3.append(["粗糙度Ra优先系列(μm)", ", ".join(str(r) for r in mech_data.ROUGHNESS_RA)])
    _style(ws3, 2)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 70

    # Sheet4 螺纹规格
    ws4 = wb.create_sheet("普通粗牙螺纹")
    ws4.append(["规格", "螺距P(mm)", "中径d2(mm)", "小径d1(mm)", "应力截面积As(mm²)"])
    for spec, p in mech_data.THREAD_M.items():
        ws4.append([spec, p["P"], p["d2"], p["d1"], p["As"]])
    _style(ws4, 5)
    for col, w in zip("ABCDE", [10, 12, 14, 14, 18]):
        ws4.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
