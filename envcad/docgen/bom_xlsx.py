# -*- coding: utf-8 -*-
"""《材料表》XLSX 自动生成（数据来自知识层）。

用法：
  from envcad.docgen.bom_xlsx import generate_material_bom
  generate_material_bom("材料表.xlsx")
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import materials


def _style_header(ws, ncols: int):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_material_bom(out_path: str, extra_items: list = None) -> str:
    wb = Workbook()

    # ── 钢筋表 ──
    ws1 = wb.active
    ws1.title = "钢筋"
    ws1.append(["序号", "公称直径 d(mm)", "牌号示例", "单根面积(mm²)",
                "单位重(kg/m)", "备注"])
    for i, (d, p) in enumerate(sorted(materials.REBAR_D.items()), 1):
        ws1.append([i, d, "HRB400", p["area"], p["w"], "常用纵筋/箍筋"])
    _style_header(ws1, 6)
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 18

    # ── 型钢表 ──
    ws2 = wb.create_sheet("型钢")
    ws2.append(["序号", "型号", "类别", "高 h(mm)", "腿宽 b(mm)",
                "截面积(cm²)", "重量(kg/m)", "截面积(mm²)"])
    idx = 1
    for cat, label, tbl in (("I", "工字钢", materials.I_BEAM),
                            ("C", "槽钢", materials.CHANNEL),
                            ("L", "等边角钢", materials.ANGLE_L),
                            ("H", "H型钢", materials.H_BEAM)):
        for name, p in tbl.items():
            ws2.append([idx, name, label, p.get("h"), p.get("b"),
                        p["A"], p["W"], round(p["Ax"], 1)])
            idx += 1
    _style_header(ws2, 8)
    for col, w in zip("ABCDEFGH", [6, 14, 12, 12, 12, 14, 12, 14]):
        ws2.column_dimensions[col].width = w

    # ── 汇总表（可选）──
    if extra_items:
        ws3 = wb.create_sheet("汇总")
        ws3.append(["序号", "名称", "规格", "数量", "单位", "备注"])
        for i, it in enumerate(extra_items, 1):
            ws3.append([i, it.get("name", ""), it.get("spec", ""),
                        it.get("qty", ""), it.get("unit", ""),
                        it.get("note", "")])
        _style_header(ws3, 6)
        for col, w in zip("ABCDEF", [6, 24, 18, 10, 8, 20]):
            ws3.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
