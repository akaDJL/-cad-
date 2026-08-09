# -*- coding: utf-8 -*-
"""液压系统设计文档自动生成（DOCX 计算书 + XLSX 元件清单）。

用法：
  from envcad.docgen.hyd_doc import generate_hyd_calc, generate_hyd_bom
  generate_hyd_calc("液压系统计算书.docx", project="XX液压站",
                    cyl=cyl, pump=pump, pipe=pipe)
  generate_hyd_bom("液压元件清单.xlsx", items=[...])
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import hyd_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_hyd_calc(out_path: str, project: str = "XX 液压系统",
                      cyl: dict = None, pump: dict = None,
                      pipe: dict = None) -> str:
    """生成液压系统设计计算书。各参数为 design.hydraulic 结果。"""
    doc = new_cn_doc(f"{project} 液压系统设计计算书")

    add_heading_cn(doc, "一、系统概述", 1)
    add_para_cn(doc, f"项目名称：{project}")
    add_para_cn(doc, "本液压系统由液压泵站、执行元件（液压缸）、控制阀组与管路组成，"
                     "采用定量泵+溢流阀调压方案。")

    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本设计依据下列现行国家标准：")
    add_table_cn(doc, ["标准编号", "名称"],
                 [[no, name] for no, name in hyd_data.hyd_code_list()])

    add_heading_cn(doc, "三、液压缸设计", 1)
    if cyl:
        add_para_cn(doc, cyl["note"])
        add_table_cn(doc, ["参数", "数值"], [
            ["工作负载 F", f"{cyl['F']} kN"],
            ["工作压力 p", f"{cyl['p']} MPa（等级 {cyl['p_grade']}）"],
            ["缸径 D", f"Φ{cyl['D']} mm"],
            ["活塞杆 d", f"Φ{cyl['d_rod']} mm"],
            ["提供推力", f"{cyl['F_actual']} kN"],
            ["流量需求 Q", f"{cyl['Q']} L/min"],
        ])

    add_heading_cn(doc, "四、液压泵选择", 1)
    if pump:
        add_para_cn(doc, pump["note"])

    add_heading_cn(doc, "五、管路管径", 1)
    if pipe:
        add_para_cn(doc, pipe["note"])
    add_para_cn(doc, "系统设置回油过滤器与空气滤清器，油箱容积按泵流量 3~5 倍选取，"
                     "并设液位液温计。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_hyd_bom(out_path: str, items: list = None) -> str:
    """生成液压元件清单。items 为 [{name, model, qty, remark}] 列表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "液压元件清单"
    ws.append(["序号", "元件名称", "型号/规格", "数量", "备注"])
    items = items or [
        dict(name="液压泵", model="定量叶片泵 PV2R", qty=1, remark="电机 7.5kW"),
        dict(name="液压缸", model="Φ100/Φ50 行程500", qty=2, remark=""),
        dict(name="溢流阀", model="DBDS-10", qty=1, remark="调压"),
        dict(name="换向阀", model="4WE10", qty=2, remark="电磁换向"),
        dict(name="过滤器", model="回油 10μm", qty=1, remark=""),
        dict(name="油箱", model="250L", qty=1, remark="带液位液温计"),
    ]
    for i, it in enumerate(items, 1):
        ws.append([i, it["name"], it.get("model", ""),
                   it.get("qty", 1), it.get("remark", "")])
    _style(ws, 5)
    for col, w in zip("ABCDE", [6, 16, 22, 8, 18]):
        ws.column_dimensions[col].width = w
    wb.save(out_path)
    return out_path
