# -*- coding: utf-8 -*-
"""化工工艺设计文档自动生成（DOCX 工艺设计说明书 + XLSX 设备管道清单）。

用法：
  from envcad.docgen.proc_doc import generate_proc_spec, generate_proc_bom
  generate_proc_spec("工艺设计说明.docx", project="XX化工装置",
                     pipe=pipe, pump=pump, hx=hx)
  generate_proc_bom("设备管道清单.xlsx", items=[...])
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import proc_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_proc_spec(out_path: str, project: str = "XX 化工装置",
                       pipe: dict = None, pump: dict = None,
                       hx: dict = None) -> str:
    """生成化工工艺设计说明书。各参数为 design.process 结果。"""
    doc = new_cn_doc(f"{project} 工艺设计说明")

    add_heading_cn(doc, "一、装置概述", 1)
    add_para_cn(doc, f"装置名称：{project}")
    add_para_cn(doc, "本装置为连续操作工艺流程，含物料输送、换热与分离单元，"
                     "按现行石油化工设计规范进行工艺、管道与设备设计。")

    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本设计依据下列现行国家标准、规范：")
    add_table_cn(doc, ["标准编号", "名称"],
                 [[no, name] for no, name in proc_data.proc_code_list()])

    add_heading_cn(doc, "三、工艺管道", 1)
    if pipe:
        add_para_cn(doc, pipe["note"])
        add_table_cn(doc, ["参数", "数值"], [
            ["流量", f"{pipe['Q']} m³/h"],
            ["经济流速", f"{pipe['v_econ']} m/s"],
            ["管径", f"DN{pipe['dn']}（内径 {pipe['di']}mm）"],
            ["实际流速", f"{pipe['v_act']} m/s"],
            ["雷诺数/流态", f"{pipe['Re']:.0f} / {pipe['flow']}"],
        ])

    add_heading_cn(doc, "四、工艺泵", 1)
    if pump:
        add_para_cn(doc, pump["note"])

    add_heading_cn(doc, "五、换热器", 1)
    if hx:
        add_para_cn(doc, hx["note"])

    add_heading_cn(doc, "六、安全与环保", 1)
    add_para_cn(doc, "设备与管道设置安全阀、爆破片等超压泄放设施；可燃气体设置检测报警；"
                     "含 VOCs 尾气经处理达标后排放，废水送污水处理装置。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_proc_bom(out_path: str, items: list = None) -> str:
    """生成设备管道清单。items 为 [{tag, name, spec, qty, remark}] 列表。"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "设备清单"
    ws1.append(["位号", "设备名称", "规格/型号", "数量", "材质", "备注"])
    items = items or [
        dict(tag="P-101", name="进料泵", spec="离心泵 Q=30m³/h H=32m",
             qty=2, mat="304", remark="一开一备"),
        dict(tag="E-101", name="换热器", spec="列管式 A=25m²",
             qty=1, mat="304", remark="逆流"),
        dict(tag="V-101", name="缓冲罐", spec="V=5m³", qty=1,
             mat="Q345R", remark=""),
        dict(tag="T-101", name="精馏塔", spec="Φ800 板式", qty=1,
             mat="304", remark=""),
    ]
    for it in items:
        ws1.append([it.get("tag", ""), it["name"], it.get("spec", ""),
                    it.get("qty", 1), it.get("mat", ""), it.get("remark", "")])
    _style(ws1, 6)
    for col, w in zip("ABCDEF", [10, 14, 24, 8, 10, 14]):
        ws1.column_dimensions[col].width = w

    ws2 = wb.create_sheet("管道规格表")
    ws2.append(["DN", "外径(mm)", "壁厚(mm)", "内径(mm)"])
    for dn, p in proc_data.PIPE_SCHEDULE.items():
        ws2.append([dn, p["od"], p["wall"], p["di"]])
    _style(ws2, 4)
    for col, w in zip("ABCD", [8, 12, 12, 12]):
        ws2.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
