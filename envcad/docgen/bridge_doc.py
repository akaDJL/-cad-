# -*- coding: utf-8 -*-
"""桥梁工程设计文档自动生成（DOCX 设计说明书 + XLSX 支座/伸缩缝选型表）。

用法：
  from envcad.docgen.bridge_doc import generate_bridge_doc, generate_bridge_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import bridge_data as bd
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_bridge_doc(out_path: str, project: str = "XX 桥梁工程",
                        lane: dict = None, bearing: dict = None,
                        joint: dict = None, girder: dict = None,
                        designer: str = "", date: str = "") -> str:
    """生成桥梁设计说明书 DOCX。"""
    doc = new_cn_doc(f"{project} 设计说明书")

    add_heading_cn(doc, "一、项目概述", 1)
    add_para_cn(doc, f"项目名称：{project}。本说明书依据 JTG D60-2015 等"
                       "公路桥梁规范编制。")
    if designer or date:
        add_para_cn(doc, f"设计：{designer}    日期：{date}")

    add_heading_cn(doc, "二、设计依据", 1)
    rows = [[no, name] for no, name in bd.bridge_code_list()]
    add_table_cn(doc, ["标准编号", "名称"], rows)

    if lane:
        add_heading_cn(doc, "三、荷载计算", 1)
        add_para_cn(doc, f"荷载等级 {lane['grade']}，跨径 {lane['L']}m，"
                         f"{lane['n_lanes']} 车道。")
        add_para_cn(doc, f"均布荷载 qk={lane['qk']} kN/m，"
                         f"集中荷载 Pk={lane['Pk']} kN，"
                         f"横向折减系数 {lane['lane_factor']}。")
        add_para_cn(doc, f"折减后总荷载：q={lane['q_total']} kN/m，"
                         f"P={lane['P_total']} kN。")

    if bearing:
        add_heading_cn(doc, "四、支座选型", 1)
        add_para_cn(doc, f"竖向力 {bearing['vertical_load']} kN，"
                         f"推荐 {bearing['bearing_spec']}，"
                         f"承载力 {bearing['capacity']} kN，"
                         f"利用率 {bearing['utilization']}%。")

    if joint:
        add_heading_cn(doc, "五、伸缩缝选型", 1)
        add_para_cn(doc, f"总伸缩量 {joint['total_displacement']} mm，"
                         f"推荐 {joint['joint_spec']}（{joint['joint_type']}）。")

    if girder:
        add_heading_cn(doc, "六、箱梁截面", 1)
        add_para_cn(doc, f"{girder['beam_type']}，跨径 {girder['span']}m，"
                         f"梁高 {girder['height']}m。")
        add_para_cn(doc, girder["note"], bold=True)

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_bridge_xlsx(out_path: str, project: str = "桥梁项目") -> str:
    """生成桥梁构件选型 XLSX。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "构件选型表"
    ws.append(["序号", "构件", "规格", "承载力/位移", "备注"])
    _style(ws, 5)

    ws2 = wb.create_sheet("板式橡胶支座(矩形)")
    ws2.append(["规格", "长×宽(mm)", "承载力(kN)", "厚度(mm)"])
    for spec, p in bd.BEARING_RECT.items():
        ws2.append([spec, f'{p["area"][0]}×{p["area"][1]}',
                    p["capacity"], p["thickness"]])
    _style(ws2, 4)

    ws3 = wb.create_sheet("伸缩缝规格")
    ws3.append(["型号", "允许位移(mm)", "类型"])
    for spec, p in bd.EXPANSION_JOINT.items():
        ws3.append([spec, p["displacement"], p["type"]])
    _style(ws3, 3)

    ws4 = wb.create_sheet("荷载等级")
    ws4.append(["等级", "qk(kN/m)", "Pk基数(kN)", "适用"])
    for grade, p in bd.ROAD_LOAD.items():
        ws4.append([grade, p["qk"], p["Pk_base"], p["note"]])
    _style(ws4, 4)

    wb.save(out_path)
    return out_path
