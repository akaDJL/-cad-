# -*- coding: utf-8 -*-
"""给排水设计文档自动生成（DOCX 设计说明书 + XLSX 用水量计算表）。

用法：
  from envcad.docgen.plumb_doc import generate_plumb_spec, generate_water_xlsx
  generate_plumb_spec("给排水设计说明.docx", project="XX给排水工程",
                      demand=demand, flow=flow, pipe=pipe, drain=drain, pump=pump)
  generate_water_xlsx("用水量计算表.xlsx", items=[...])
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..knowledge import plumb_data
from . import new_cn_doc, add_heading_cn, add_para_cn, add_table_cn


def generate_plumb_spec(out_path: str, project: str = "XX 给排水工程",
                        demand: dict = None, flow: dict = None,
                        pipe: dict = None, drain: dict = None,
                        pump: dict = None) -> str:
    """生成给排水设计说明书。各参数为 design.plumbing 结果。"""
    doc = new_cn_doc(f"{project} 给排水设计说明")

    add_heading_cn(doc, "一、工程概况", 1)
    add_para_cn(doc, f"工程名称：{project}")
    add_para_cn(doc, "本工程包括生活给水、生活排水与雨水系统设计，"
                     "市政给水直供低区，高区采用变频加压供水。")

    add_heading_cn(doc, "二、设计依据", 1)
    add_para_cn(doc, "本工程依据下列现行国家标准、规范：")
    add_table_cn(doc, ["标准编号", "名称"],
                 [[no, name] for no, name in plumb_data.plumb_code_list()])

    add_heading_cn(doc, "三、生活用水量", 1)
    if demand:
        add_para_cn(doc, demand["note"])
        add_table_cn(doc, ["项目", "数值"], [
            ["用水定额", f"{demand['quota']} {demand['unit']}"],
            ["最高日用水量", f"{demand['Qd']} m³/d"],
            ["最大时用水量", f"{demand['Qh_max']} m³/h"],
            ["最大时秒流量", f"{demand['Qs']} L/s"],
        ])

    add_heading_cn(doc, "四、给水系统", 1)
    if flow:
        add_para_cn(doc, "设计秒流量：" + flow["note"])
    if pipe:
        add_para_cn(doc, "给水管径：" + pipe["note"])

    add_heading_cn(doc, "五、排水系统", 1)
    if drain:
        add_para_cn(doc, drain["note"])
    add_para_cn(doc, "室内采用污废合流，设置伸顶通气管；室外雨污分流排入市政管网。")

    add_heading_cn(doc, "六、加压设备", 1)
    if pump:
        add_para_cn(doc, "给水泵扬程：" + pump["note"])
    add_para_cn(doc, "消防给水按现行《消防给水及消火栓系统技术规范》GB 50974 单独设计。")

    doc.save(out_path)
    return out_path


def _style(ws, ncols):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="黑体", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


def generate_water_xlsx(out_path: str, items: list = None) -> str:
    """生成用水量计算表。items 为 [{name, number, kind}] 列表。"""
    from ..design.plumbing import design_water_demand
    wb = Workbook()
    ws = wb.active
    ws.title = "用水量计算表"
    ws.append(["序号", "用水部位", "用水单位数", "定额", "单位",
               "最高日(m³/d)", "最大时(m³/h)", "秒流量(L/s)"])
    items = items or [
        dict(name="办公楼", number=500, kind="办公楼"),
        dict(name="职工宿舍", number=300, kind="宿舍"),
        dict(name="食堂", number=800, kind="餐饮"),
    ]
    tot_d = 0.0
    for i, it in enumerate(items, 1):
        try:
            r = design_water_demand(it["number"], it.get("kind", "办公楼"))
        except Exception as _e:
            continue
        ws.append([i, it["name"], r["number"], r["quota"], r["unit"],
                   r["Qd"], r["Qh_max"], r["Qs"]])
        tot_d += r["Qd"]
    ws.append(["", "合计", "", "", "", round(tot_d, 1), "", ""])
    _style(ws, 8)
    for col, w in zip("ABCDEFGH", [6, 14, 12, 16, 10, 13, 13, 12]):
        ws.column_dimensions[col].width = w
    wb.save(out_path)
    return out_path
